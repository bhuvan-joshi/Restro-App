import openpyxl
import os
from openpyxl.drawing.image import Image
import pandas as pd

# Configuration
EXCEL_FILE = "PL- ARPER.xlsx"

def check_excel_file():
    """Check the Excel file for images and structure"""
    print(f"Checking Excel file: {EXCEL_FILE}")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file '{EXCEL_FILE}' not found!")
        return
    
    # Load the workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Process each worksheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {sheet.dimensions}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        
        # Check for images
        if hasattr(sheet, '_images'):
            print(f"Images in sheet: {len(sheet._images)}")
            for i, img in enumerate(sheet._images):
                if hasattr(img, 'anchor'):
                    anchor = img.anchor
                    if hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                        print(f"  Image {i+1}: at row {anchor.row}, col {anchor.col}")
                    else:
                        print(f"  Image {i+1}: anchor doesn't have row/col attributes")
                else:
                    print(f"  Image {i+1}: no anchor information")
        else:
            print("No _images attribute found in sheet")
            
        # Try alternative methods to find images
        print("Checking for drawings...")
        if hasattr(sheet, '_drawing'):
            print(f"  Drawing found: {sheet._drawing}")
            if hasattr(sheet._drawing, '_shapes'):
                print(f"  Shapes in drawing: {len(sheet._drawing._shapes)}")
                for i, shape in enumerate(sheet._drawing._shapes):
                    print(f"    Shape {i+1}: {type(shape)}")
        else:
            print("  No _drawing attribute found")
            
        # Try to read as pandas dataframe to see structure
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            print(f"DataFrame shape: {df.shape}")
            print("Column names:")
            for col in df.columns:
                print(f"  {col}")
        except Exception as e:
            print(f"Error reading as DataFrame: {e}")

if __name__ == "__main__":
    check_excel_file()
